import time
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

def search(search_terms):
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get('https://www.google.com/maps')
    driver.maximize_window()
    time.sleep(3)

    all_links = {}

    for search_value in search_terms:
        try:
            search_box = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.ID, 'searchboxinput')))
            search_box.clear()
            search_box.send_keys(search_value)
            search_box.send_keys(Keys.RETURN)
        except Exception as e:
            print(f"Error finding search box for {search_value}: {e}")
            continue

        try:
            scrollable_div = WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//div[@class='m6QErb DxyBCb kA9KIf dS8AEf XiKgde ecceSd']")))
        except Exception as e:
            print(f"Error finding scrollable div for {search_value}: {e}")
            continue

        previous_length = 0

        while True:
            scrollable_div.send_keys(Keys.END)
            time.sleep(2)
            new_results = driver.find_elements(By.CLASS_NAME, 'Nv2PK')

            if len(new_results) == previous_length:
                break
            previous_length = len(new_results)

        links = []

        elements = driver.find_elements(By.CSS_SELECTOR, 'a.hfpxzc')
        for element in elements:
            link = element.get_attribute('href')
            links.append(link)

        all_links[search_value] = links

    driver.quit()


    file_name = 'search_results.xlsx'

    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
    else:
        workbook = openpyxl.load_workbook(file_name)

    for search_value, links in all_links.items():
        if search_value in workbook.sheetnames:
            sheet = workbook[search_value]
        else:
            sheet = workbook.create_sheet(title=search_value)

        for row_index, link in enumerate(links, start=1):
            sheet.cell(row=row_index, column=2, value=link)  # Write link in the second column

    workbook.save(file_name)

    print("Links have been saved to 'search_results.xlsx'.")


with open('map_search.txt', 'r') as file:
    search_terms = [each_row.strip() for each_row in file]
    for row in file:
        search_terms = row.strip()
search(search_terms)

