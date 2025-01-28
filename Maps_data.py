#import csv
import time
import re
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

def search(url, sheet_name):
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.get(url)
    driver.maximize_window()
    time.sleep(3)

    link = f"URL :{url}"
    print(link)
    data = []
    data.append(link)

    try:
        title_element = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".DUwDvf.lfPIob")))
        title = title_element.text.strip()
        Main_title = f"Title : {title}"
        data.append(Main_title)
        print(Main_title)
    except Exception as error:
        print((f"Title not found: {error}"))

    try:
        all_rating = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".F7nice span[aria-hidden='true']")))
        over_all_rating = f"Rating :{all_rating.text}"
        data.append(over_all_rating)
        print(over_all_rating)
    except Exception as rating_error:
        print(f"Rating Not Found: {rating_error}")

    try:
        over_all_reviews =  WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".F7nice span[aria-label]")))
        for elements in over_all_reviews:
            if '(' in elements.text and ')' in elements.text:
                reviews_counting = f"Total Reviews : {elements.text.strip('()')}"
                data.append(reviews_counting)
                print(reviews_counting)
                break
    except Exception as e:
        print((f"Total Reviews Not Found: {e}"))

    try:
        website = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".rogA2c.ITvuef .Io6YTe.fontBodyMedium.kR99db.fdkmkc")))
        website_link = f"Website :{website.text.strip()}"
        data.append(website_link)
        print(website_link)
    except Exception as E:
        print(f"Website Link Not Found :{E}")

    try:
        Number = ''
        all_elements =  WebDriverWait(driver,20).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".Io6YTe.fontBodyMedium.kR99db.fdkmkc")))
        country_code_pattern = re.compile(r"\+\d{1,3}")
        for phone in all_elements:
            text = phone.text.strip()
            if country_code_pattern.search(text):
                for char in text:
                    if char != ' ':
                        Number = Number + char
                phone_Number = f"Phone Number: {Number}"
                data.append(phone_Number)
                print(phone_Number)
                break
    except Exception as e:
        print(f"Phone Number Not Found: {e}")

    try:
        reviews_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "button[aria-label*='Reviews for']")))
        reviews_button.click()
        time.sleep(10)

        scroll = WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".m6QErb.DxyBCb.kA9KIf.dS8AEf.XiKgde ")))

        previous_length = 0
        same_length_count = 0
        max_same_length = 2

        while True:
            driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", scroll)
            time.sleep(3)
            new_results = driver.find_elements(By.CSS_SELECTOR, ".jftiEf.fontBodyMedium ")

            if len(new_results) == previous_length:
                same_length_count += 1
            else:
                same_length_count = 0  # Reset counter if new reviews are loaded

            previous_length = len(new_results)  # Update previous_length with the new length

            if same_length_count >= max_same_length:
                all_results = driver.find_elements(By.CSS_SELECTOR, ".jftiEf.fontBodyMedium ")
                for each_review in all_results:
                    try:
                        star = 0
                        name_element = each_review.find_element(By.CSS_SELECTOR, ".d4r55")
                        comment_element = each_review.find_element(By.CSS_SELECTOR, ".wiI7pd")
                        filled_stars = each_review.find_elements(By.CSS_SELECTOR, ".elGi1d")
                        star = len(filled_stars)
                        combined = f"Review \nName : {name_element.text.strip()} \nStars : {star} \nComment : {comment_element.text.strip()}"
                        data.append(combined)
                        print(combined)
                    except Exception as review_error:
                        print(f"Error processing review: {review_error}")
                break
    except Exception as e:
        print(f"Button Not found: {e}")
    finally:
        driver.quit()

    filename = file_name = 'Output.xlsx'

    # file_exists = os.path.isfile(filename)
    # with open(filename, mode='a' if file_exists else 'w', newline='',encoding='utf-8') as file:
    #     writer = csv.writer(file)
    #     writer.writerow(data)

    if not os.path.exists(file_name):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
    else:
        workbook = openpyxl.load_workbook(file_name)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

    next_row = sheet.max_row + 1
    for index, value in enumerate(data, start=1):
        sheet.cell(row=next_row, column=index, value=value)

    workbook.save(file_name)



# website = 'https://www.google.com/maps/place/Fauji+Foundation+Hospital/@31.4893287,74.3681403,14z/data=!4m16!1m9!3m8!1s0x391905f289ed65bd:0xc08fcb8d061b585b!2sFauji+Foundation+Hospital!8m2!3d31.4997347!4d74.3947478!9m1!1b1!16s%2Fm%2F0qs9_cx!3m5!1s0x391905f289ed65bd:0xc08fcb8d061b585b!8m2!3d31.4997347!4d74.3947478!16s%2Fm%2F0qs9_cx?authuser=0&hl=en&entry=ttu&g_ep=EgoyMDI1MDEyMi4wIKXMDSoASAFQAw%3D%3D'
# search(website)


def links():

    file_name = "search_results.xlsx"
    workbook = openpyxl.load_workbook(file_name)

    for sheet_name in workbook.sheetnames:
#        print(f"Sheet Name: {sheet_name}")
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
            first_cell = row[0]
            second_cell = row[1]


            if first_cell.value != "Done":
                if first_cell.value != "Done" and second_cell.value:
                    search(second_cell.value, sheet_name)
                    first_cell.value = "Done"
                    workbook.save(file_name)

    workbook.close()

links()
